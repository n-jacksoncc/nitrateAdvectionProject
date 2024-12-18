import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

"""
This code is meant to numerically model the 1d advection of nitrate through the Arkansas River. It utilizes 
available USGS data for data collection site discharge within the river and potential nitrate concentration contributors. 
This code can be adapted for other rivers. 

Input Parameters:
    wb = the file path unique to the user leading to the Excel spreadsheet with velocity values
    u = velocity values (m/s) currently assuming that velocity values will be input in the E column in excel
    dx = the change in x between nodes (m) 
    x = the grid of distances from the start location (m)
    C = the initial concentration of nitrate (mg/L) at the initial source location
    dt = the change in time at which each calculation will be performed (s)
    endtime =  the total amount of time the model will run for (s)
    city_locations = optional input list of the distance at which cities neighboring the river are located ([m])
    city_names = optional input list of the names of neighboring cities ([strings])
  
    
Output Parameters:
    courant = the stability parameter for advection --> ensure that this value is as close to, but does not exceed 1
    one plots: river velocity over space and advection of nitrate over space initially and at the end time
"""



'SET UP INITIAL CONDITIONS' 

# set up the workbook and worksheet for loading in the data
wb = load_workbook("/Users/nadia/Downloads/NumericalModeling/Homework/finalProject/velocity_data.xlsx")
ws = wb.active

# the dx is currently set to 1000 to increase stability due to the variety in distances between the velocity measurements
dx = 1000 # meters

# the distance from USGS site #07091200 to the border between Colorado and Kansas (measured via GoogleEarth)
x = np.arange(0, 363000, dx)  #meters
nodes = len(x)

#velocity in m/s
u = np.zeros(nodes)
u[(x<60433)] = ws['E2'].value #site number 7091200
u[(60433<=x) & (x<133863)] = ws['E3'].value  #site number 7094500
u[(133863<=x) & (x<150633)] = ws['E4'].value  # site number 7099973
u[(150633<=x) & (x<253793)] =  ws['E5'].value  # site number 7109500
u[(253793<=x) & (x<279563)] =  ws['E6'].value  # site number 71224000
u[(279563<=x) & (x<310262)] =  ws['E7'].value  # site number 7130500
u[(310262<=x) & (x<336549)] =  ws['E8'].value  # site number 7133000
u[(336549<=x) & (x<354592)] =  ws['E9'].value  # site number 7134180
u[(354592<x)] = ws['E10'].value  # site number 7135500


# the initial concentration of nitrate between 0 and 30.216 km
C = np.zeros(nodes)
C[(x<=30216)] = 100 # mg/L


'SET UP THE A MATRIX'

# u is not constant, but courant must be less than one
dt = dx/np.max(u) # seconds

courant = dt*u/dx

A = np.zeros((nodes, nodes))

for i in range(1,nodes-1):
    if i > 0:
        A[i,i]= 1-courant[i]
        A[i,i-1]= courant[i]
        
        
        
'PLOT THE INITIAL CONDITIONS'

# set up the figure for plotting concentration
fig, ax2 = plt.subplots(figsize=(10, 6))

# create a second y-axis for velocity
ax1 = ax2.twinx()

# plot the velocity and fill the area under it
line1, = ax1.plot(x, u, label="Measured River Velocity (m/s)", color='cornflowerblue', alpha = 0.2)
ax1.fill_between(x, u, color='cornflowerblue', alpha = 0.175)


# list the neighboring cities and their respective locations 
city_locations = [60433, 150633, 253793, 310762]
city_names = ['Parkdale', 'Pueblo', 'Las Animas', "Lamar"]


# annotate the plot with the city names
for city, loc in zip(city_names, city_locations):
    ax1.annotate(
        city, 
        xy=(loc, 0),  # the point to annotate (x, y)
        xytext=(loc, 0.15),  # the location of the text (height from x-axis)
        textcoords='data',  
        ha='center',  # center the text
        va='bottom',  # place text just above the point
        fontsize=10, 
        color='darkgreen',
        arrowprops=dict(arrowstyle='->', lw =1, color='darkgreen')  # arrow pointing to the city location
    )


# Set labels for the right y-axis (velocity)
ax1.set_ylabel("River Velocity (m/s)", color='cornflowerblue')
ax1.tick_params(axis='y', labelcolor='cornflowerblue')

# plot the initial nitrate concentration
line2, = ax2.plot(x, C,'--k',linewidth = 3, label="Initial Nitrate Concentration (mg/L)", color='k')



' DOT PRODUCT'

seconds = 0
endtime = 60*60*24  #1 day

# dot product of the A matrix and the concentration matrix
while seconds <= endtime:
    C_new = np.dot(A, C)
    C[:] = C_new*1
    seconds += dt
   
    
'PLOT THE FINAL CONDITIONS'

# plot the final nitrate concentration
line3, = ax2.plot(x, C, linewidth = 3, label = "Nitrate Concentration After 24 hours", color = "gray")
ax2.set_xlabel("Distance from Initial Data Collection Site (m)")
ax2.set_ylabel("Nitrate Concentration (mg/L)")
ax2.set_title("Nitrate Advection Through the Arkansas River") 

# adjust the axis for aesthetics
ax1.set_ylim(0, 2.5)
ax2.set_ylim(0, 115)
ax1.set_xlim(0, 360000)

# had to do something funky because legend was not cooperating
# keep track of labels and lines 
handles = [line1, line2, line3]  
labels = [line1.get_label(), line2.get_label(), line3.get_label()]  
# add the legend to the plot
ax2.legend(handles, labels)

fig.tight_layout()
plt.show()